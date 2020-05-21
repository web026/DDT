from geopy.geocoders import Nominatim
from openpyxl import load_workbook
from nose.tools import *
import unittest
from decimal import Decimal

delta = 0.98


class DataTest(unittest.TestCase):

    def test(self):
        wb = load_workbook('./data_coord.xlsx')

        sheet = wb['Лист1']
        for k in range(1,sheet.max_row):
            cellA = sheet.cell(row=k, column=1).value
            cellB = Decimal(str(sheet.cell(row=k, column=2).value))
            cellC = Decimal(str(sheet.cell(row=k, column=3).value))
            geol = Nominatim(user_agent="my-app")
            locate = geol.geocode(cellA)
            latitude = Decimal(str(locate.latitude))
            longitude = Decimal(str(locate.longitude))
            if latitude > cellB:
                diff_1 = cellB / latitude
            else:
                diff_1 = latitude / cellB
            if longitude > cellC:
                diff_2 = cellC / longitude
            else:
                diff_2 = longitude / cellC

            diff = (diff_1+diff_2)/2*100
            try:
                self.assertGreaterEqual(diff, delta, msg=None)
                diff = diff.quantize(Decimal("3.00"))
                print('Совпадение на ' + str(diff) + '%, тест пройден')
            except AssertionError as e:
                print("Raise {}".format(e))

#

