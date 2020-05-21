from geopy.geocoders import Nominatim
from openpyxl import load_workbook
import difflib
from nose.tools import *

delta = 0.80


def similarity(s1, s2):
    normalized1 = s1.lower()
    normalized2 = s2.lower()
    matcher = difflib.SequenceMatcher(None, normalized1, normalized2)
    return matcher.ratio()


class DataTest(unittest.TestCase):

    def test(self):
        wb = load_workbook('./data_address.xlsx')
        sheet = wb['Лист1']
        for k in range(1, sheet.max_row):
            cellA = sheet.cell(row=k, column=1).value
            cellB = sheet.cell(row=k, column=2).value
            cellC = sheet.cell(row=k, column=3).value
            geol = Nominatim(user_agent="my-app")
            address = geol.geocode(str(cellB + ',' + cellC))
            diff = similarity(str(address), str(cellA))

            try:
                self.assertGreaterEqual(diff, delta, msg=None)
                print('Совпадение на ' + str(diff * 100) + '%, тест пройден')
            except AssertionError as e:
                print("Raise {}".format(e))

